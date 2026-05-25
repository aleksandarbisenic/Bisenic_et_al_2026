#!/usr/bin/env python3
"""
extract_go_hits_adjacent_columns.py

Goal:
- Read multiple genome annotation XLSX files from a folder
- For each GO term in go_terms.txt, output one block in a single CSV file
- Genomes appear as adjacent column pairs: (Gene_ID, Description)

Block format per GO term:
Row 1: GO term in first cell, rest empty
Row 2: headers for each genome pair: Genome.xlsx (Gene_ID), Genome.xlsx (Description)
Row 3+: gene rows aligned across genomes (blank if shorter)
Then one blank row, next GO term block

XLSX expectations (same logic as your old script):
- Column A: Gene ID
- Column B: row type (e.g., 'original_DE')
- Column F: either GO term OR description (when row type == 'original_DE')
- GO term matching is exact (strip whitespace and match full string)
"""

from __future__ import annotations

import argparse
import csv
import os
from typing import Dict, List, Optional, Set, Tuple

import openpyxl


def read_go_terms(go_txt_path: str) -> List[str]:
    """One GO term per line, ignores blank lines and # comments, preserves order, de-duplicates."""
    terms: List[str] = []
    seen: Set[str] = set()
    with open(go_txt_path, "r", encoding="utf-8") as f:
        for raw in f:
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            if line not in seen:
                seen.add(line)
                terms.append(line)
    return terms


def find_xlsx_files(folder: str) -> List[str]:
    """Sorted list of .xlsx files; skips Excel temp lockfiles."""
    files: List[str] = []
    for fn in os.listdir(folder):
        if fn.lower().endswith(".xlsx") and not fn.startswith("~$"):
            files.append(os.path.join(folder, fn))
    files.sort()
    return files


def extract_hits_from_xlsx(
    xlsx_path: str,
    term_set: Set[str],
) -> Tuple[Dict[str, List[str]], Dict[str, str]]:
    """
    Returns:
      term_to_gene_ids: GO_term -> list of unique Gene_IDs (in first-seen order)
      gene_to_desc: Gene_ID -> Description (captured from 'original_DE' rows)
    """
    term_to_gene_ids: Dict[str, List[str]] = {}
    term_seen: Dict[str, Set[str]] = {}
    gene_to_desc: Dict[str, str] = {}

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    current_gene_id: Optional[str] = None
    current_desc: str = ""

    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 6:
            continue

        gene_id_val = row[0]   # col A
        row_type = row[1]      # col B
        col_f = row[5]         # col F

        # Update current gene id
        if gene_id_val is not None:
            gid = str(gene_id_val).strip()
            if gid:
                if gid != current_gene_id:
                    current_gene_id = gid
                    current_desc = gene_to_desc.get(gid, "")
        else:
            gid = current_gene_id

        if gid is None:
            continue

        # Capture description from original_DE rows
        if isinstance(row_type, str) and row_type.strip() == "original_DE":
            if col_f is not None:
                current_desc = str(col_f).strip()
                gene_to_desc[gid] = current_desc

        # Exact GO-term match in column F
        if col_f is None:
            continue
        fval = str(col_f).strip()

        if fval in term_set:
            term = fval
            if term not in term_to_gene_ids:
                term_to_gene_ids[term] = []
                term_seen[term] = set()

            if gid not in term_seen[term]:
                term_seen[term].add(gid)
                term_to_gene_ids[term].append(gid)

                # store current description if available
                if gid not in gene_to_desc and current_desc:
                    gene_to_desc[gid] = current_desc

    wb.close()
    return term_to_gene_ids, gene_to_desc


def main() -> None:
    ap = argparse.ArgumentParser(
        description="Extract GO-term gene hits from many XLSX genomes into ONE CSV with adjacent genome columns."
    )
    ap.add_argument("xlsx_folder", help="Folder containing genome annotation XLSX files")
    ap.add_argument("go_terms_txt", help="Text file with GO terms (one per line; exact match)")
    ap.add_argument("-o", "--out", default="go_hits_adjacent_columns.csv", help="Output CSV filename")
    args = ap.parse_args()

    if not os.path.isdir(args.xlsx_folder):
        raise SystemExit(f"ERROR: folder not found: {args.xlsx_folder}")

    go_terms = read_go_terms(args.go_terms_txt)
    if not go_terms:
        raise SystemExit("ERROR: go_terms.txt is empty (or only comments/blank lines).")

    xlsx_paths = find_xlsx_files(args.xlsx_folder)
    if not xlsx_paths:
        raise SystemExit(f"ERROR: no .xlsx files found in folder: {args.xlsx_folder}")

    # Genome names are EXACT filenames including extension (your requirement)
    genomes: List[str] = [os.path.basename(p) for p in xlsx_paths]
    term_set = set(go_terms)

    # Data containers:
    # hits[genome][term] = [gene_id1, gene_id2, ...]
    hits: Dict[str, Dict[str, List[str]]] = {g: {} for g in genomes}
    # descs[genome][gene_id] = description
    descs: Dict[str, Dict[str, str]] = {g: {} for g in genomes}

    # Parse all xlsx
    for xlsx_path, genome in zip(xlsx_paths, genomes):
        term_to_gene_ids, gene_to_desc = extract_hits_from_xlsx(xlsx_path, term_set)
        hits[genome] = term_to_gene_ids
        descs[genome] = gene_to_desc

        total_hits = sum(len(v) for v in term_to_gene_ids.values())
        print(f"[OK] {genome}: {total_hits} total GO hits")

    # Write one CSV with GO-term blocks
    with open(args.out, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)

        for term in go_terms:
            # -------- Row 1: GO term label row --------
            # First cell has term, rest blank (2 columns per genome)
            w.writerow([term] + [""] * (2 * len(genomes) - 1))

            # -------- Row 2: genome headers --------
            header_row = []
            for g in genomes:
                header_row.extend([f"{g} | Gene_ID", f"{g} | Description"])
            w.writerow(header_row)

            # -------- Data rows --------
            # Determine max gene count across genomes for this term
            max_len = 0
            per_genome_gene_lists: Dict[str, List[str]] = {}
            for g in genomes:
                gene_list = hits.get(g, {}).get(term, [])
                per_genome_gene_lists[g] = gene_list
                if len(gene_list) > max_len:
                    max_len = len(gene_list)

            # If no hits in any genome, still leave an empty row for readability
            if max_len == 0:
                w.writerow([""] * (2 * len(genomes)))
                w.writerow([])  # blank line between GO terms
                continue

            # Fill block rows
            for i in range(max_len):
                row_out: List[str] = []
                for g in genomes:
                    gene_list = per_genome_gene_lists[g]
                    if i < len(gene_list):
                        gid = gene_list[i]
                        row_out.append(gid)
                        row_out.append(descs.get(g, {}).get(gid, ""))
                    else:
                        row_out.extend(["", ""])
                w.writerow(row_out)

            # -------- Blank row between GO terms --------
            w.writerow([])

    print(f"\nDone. Output written to: {args.out}")


if __name__ == "__main__":
    main()
