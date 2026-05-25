#!/usr/bin/env python3
"""
level3_gene_search.py

Multi-genome "contains" search across annotation XLSX files.

DEFAULT BEHAVIOR:
✅ Case-sensitive substring matching

Optional:
--ignore-case   -> makes matching case-insensitive

- Reads all .xlsx files in a folder (one per genome)
- Reads search terms from a text file (one term per line)
- For each term:
    - Writes a block into ONE CSV output
    - Genomes appear as adjacent 2-column pairs: Gene_ID + Description
    - Rows are gene hits (max hits across genomes determines block height)
- Matching rule: Column F contains the term (substring match)
- Description is only trusted from rows where column B == "original_DE"
"""

from __future__ import annotations

import argparse
import csv
import os
from typing import Dict, List, Optional, Set, Tuple

import openpyxl


def read_terms(terms_txt_path: str) -> List[str]:
    """One term per line, ignore blanks and # comments, preserve order, de-duplicate."""
    terms: List[str] = []
    seen: Set[str] = set()
    with open(terms_txt_path, "r", encoding="utf-8") as f:
        for raw in f:
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            if line not in seen:
                seen.add(line)
                terms.append(line)
    return terms


def find_xlsx_files(folder: str) -> List[str]:
    """Sorted list of .xlsx files; skip Excel temp lock files (~$...)."""
    files: List[str] = []
    for fn in os.listdir(folder):
        if fn.lower().endswith(".xlsx") and not fn.startswith("~$"):
            files.append(os.path.join(folder, fn))
    files.sort()
    return files


def extract_hits_from_xlsx_contains(
    xlsx_path: str,
    terms: List[str],
    ignore_case: bool = False,  # ✅ default is CASE-SENSITIVE now
) -> Tuple[Dict[str, List[str]], Dict[str, str]]:
    """
    Returns:
      term_to_gene_ids: term -> list of unique Gene_IDs (first-seen order)
      gene_to_desc: Gene_ID -> Description (captured ONLY from rows where B == 'original_DE')
    """
    term_to_gene_ids: Dict[str, List[str]] = {}
    term_seen: Dict[str, Set[str]] = {}
    gene_to_desc: Dict[str, str] = {}

    # Prepare terms
    if ignore_case:
        term_pairs = [(t, t.lower()) for t in terms]
    else:
        term_pairs = [(t, t) for t in terms]

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    current_gene_id: Optional[str] = None
    current_desc: str = ""

    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 6:
            continue

        gene_id_val = row[0]   # A
        row_type = row[1]      # B
        col_f = row[5]         # F

        # Update current gene ID
        if gene_id_val is not None:
            gid = str(gene_id_val).strip()
            if gid:
                if gid != current_gene_id:
                    current_gene_id = gid
                    current_desc = gene_to_desc.get(gid, "")
        else:
            gid = current_gene_id

        if gid is None:
            continue

        # Capture description ONLY from "original_DE" rows
        if isinstance(row_type, str) and row_type.strip() == "original_DE":
            if col_f is not None:
                current_desc = str(col_f).strip()
                gene_to_desc[gid] = current_desc

        # Contains match
        if col_f is None:
            continue
        fval = str(col_f).strip()
        if not fval:
            continue

        haystack = fval.lower() if ignore_case else fval

        for original_term, term_key in term_pairs:
            if term_key in haystack:
                if original_term not in term_to_gene_ids:
                    term_to_gene_ids[original_term] = []
                    term_seen[original_term] = set()

                if gid not in term_seen[original_term]:
                    term_seen[original_term].add(gid)
                    term_to_gene_ids[original_term].append(gid)

                    # Store description if we already have it in-hand
                    if gid not in gene_to_desc and current_desc:
                        gene_to_desc[gid] = current_desc

    wb.close()
    return term_to_gene_ids, gene_to_desc


def main() -> None:
    ap = argparse.ArgumentParser(
        description="Extract substring/contains matches from multiple XLSX genome annotation files into ONE CSV."
    )
    ap.add_argument("xlsx_folder", help="Folder containing XLSX files (one per genome)")
    ap.add_argument("terms_txt", help="Text file with search terms (one per line)")
    ap.add_argument("-o", "--out", default="term_hits_contains.csv", help="Output CSV filename")

    # ✅ NEW flag (instead of case-sensitive flag)
    ap.add_argument(
        "--ignore-case",
        action="store_true",
        help="If set, matching becomes case-insensitive (default is case-sensitive).",
    )

    args = ap.parse_args()

    if not os.path.isdir(args.xlsx_folder):
        raise SystemExit(f"ERROR: folder not found: {args.xlsx_folder}")

    terms = read_terms(args.terms_txt)
    if not terms:
        raise SystemExit("ERROR: terms file is empty (or only comments/blanks).")

    xlsx_paths = find_xlsx_files(args.xlsx_folder)
    if not xlsx_paths:
        raise SystemExit(f"ERROR: no .xlsx files found in folder: {args.xlsx_folder}")

    genomes: List[str] = [os.path.basename(p) for p in xlsx_paths]
    ignore_case = args.ignore_case  # ✅ default False

    hits: Dict[str, Dict[str, List[str]]] = {}
    descs: Dict[str, Dict[str, str]] = {}

    for xlsx_path, genome in zip(xlsx_paths, genomes):
        term_to_gene_ids, gene_to_desc = extract_hits_from_xlsx_contains(
            xlsx_path=xlsx_path,
            terms=terms,
            ignore_case=ignore_case,
        )
        hits[genome] = term_to_gene_ids
        descs[genome] = gene_to_desc

        total_hits = sum(len(v) for v in term_to_gene_ids.values())
        print(f"[OK] {genome}: {total_hits} total hits")

    with open(args.out, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)

        for term in terms:
            # Term title row
            w.writerow([term] + [""] * (2 * len(genomes) - 1))

            # Headers
            header_row: List[str] = []
            for g in genomes:
                header_row.extend([f"{g} | Gene_ID", f"{g} | Description"])
            w.writerow(header_row)

            # Lists per genome for this term
            per_genome_gene_lists: Dict[str, List[str]] = {}
            max_len = 0
            for g in genomes:
                gene_list = hits.get(g, {}).get(term, [])
                per_genome_gene_lists[g] = gene_list
                if len(gene_list) > max_len:
                    max_len = len(gene_list)

            if max_len == 0:
                w.writerow([""] * (2 * len(genomes)))
                w.writerow([])
                continue

            for i in range(max_len):
                row_out: List[str] = []
                for g in genomes:
                    gene_list = per_genome_gene_lists[g]
                    if i < len(gene_list):
                        gid = gene_list[i]
                        row_out.append(gid)
                        row_out.append(descs.get(g, {}).get(gid, ""))
                    else:
                        row_out.extend(["", ""])
                w.writerow(row_out)

            # Blank row separator
            w.writerow([])

    print(f"\nDone. Output written to: {args.out}")


if __name__ == "__main__":
    main()
