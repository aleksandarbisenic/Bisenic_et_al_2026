#!/usr/bin/env python3
"""
extract_hits_contains_originalDE_adjacent_columns.py

What it does:
- Reads multiple genome annotation XLSX files from a folder
- Reads search terms from terms.txt (one per line)
- Records hits only when:
    Column B == "original_DE"
  AND Column F contains the search term (substring match)
- Outputs ONE CSV file organized term-by-term
- Within each term block: genomes are adjacent column pairs (Gene_ID + Description)

Block format per term:
Row 1: term in first cell, rest empty
Row 2: genome headers: [Genome.xlsx | Gene_ID, Genome.xlsx | Description] repeated
Row 3+: hits aligned row-wise per genome
Then one blank row, next term block

Important:
- Matching is case-insensitive by default.
- This script searches ONLY the gene description rows (original_DE),
  which is ideal for targeted “find these genes even if GO labels are missing.”
"""

from __future__ import annotations

import argparse
import csv
import os
from typing import Dict, List, Optional, Set, Tuple

import openpyxl


def read_terms(term_txt_path: str) -> List[str]:
    """One search term per line. Ignores blank lines and # comments. Preserves order. De-duplicates."""
    terms: List[str] = []
    seen: Set[str] = set()
    with open(term_txt_path, "r", encoding="utf-8") as f:
        for raw in f:
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            if line not in seen:
                seen.add(line)
                terms.append(line)
    return terms


def find_xlsx_files(folder: str) -> List[str]:
    """Sorted list of .xlsx files; skips Excel temp lockfiles."""
    files: List[str] = []
    for fn in os.listdir(folder):
        if fn.lower().endswith(".xlsx") and not fn.startswith("~$"):
            files.append(os.path.join(folder, fn))
    files.sort()
    return files


def extract_hits_from_xlsx_contains_originalDE(
    xlsx_path: str,
    terms: List[str],
    case_insensitive: bool = True,
) -> Tuple[Dict[str, List[str]], Dict[str, str]]:
    """
    Only searches within rows where Column B == 'original_DE'.

    Returns:
      term_to_gene_ids: term -> list of unique Gene_IDs (ordered by appearance)
      gene_to_desc: gene_id -> description (from original_DE rows)
    """

    if case_insensitive:
        term_patterns = [(t, t.lower()) for t in terms]
    else:
        term_patterns = [(t, t) for t in terms]

    term_to_gene_ids: Dict[str, List[str]] = {t: [] for t in terms}
    term_seen: Dict[str, Set[str]] = {t: set() for t in terms}
    gene_to_desc: Dict[str, str] = {}

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    current_gene_id: Optional[str] = None

    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 6:
            continue

        gene_id_val = row[0]   # col A
        row_type = row[1]      # col B
        col_f = row[5]         # col F (description on original_DE rows)

        # Update current Gene_ID
        if gene_id_val is not None:
            gid = str(gene_id_val).strip()
            if gid:
                current_gene_id = gid
        else:
            gid = current_gene_id

        if gid is None:
            continue

        # We ONLY care about original_DE rows
        if not (isinstance(row_type, str) and row_type.strip() == "original_DE"):
            continue

        if col_f is None:
            continue

        desc = str(col_f).strip()
        if not desc:
            continue

        # Store description (canonical)
        gene_to_desc[gid] = desc

        # Contains matching within original_DE descriptions only
        desc_cmp = desc.lower() if case_insensitive else desc

        for original_term, term_cmp in term_patterns:
            if term_cmp in desc_cmp:
                if gid not in term_seen[original_term]:
                    term_seen[original_term].add(gid)
                    term_to_gene_ids[original_term].append(gid)

    wb.close()
    return term_to_gene_ids, gene_to_desc


def main() -> None:
    ap = argparse.ArgumentParser(
        description="Extract CONTAINS-matched hits from original_DE gene description rows (many XLSX → one CSV)."
    )
    ap.add_argument("xlsx_folder", help="Folder containing genome annotation XLSX files")
    ap.add_argument("terms_txt", help="Text file with search terms (one per line)")
    ap.add_argument("-o", "--out", default="term_hits_contains_originalDE.csv", help="Output CSV filename")
    ap.add_argument(
        "--case-sensitive",
        action="store_true",
        help="Use case-sensitive matching (default: case-insensitive)",
    )
    args = ap.parse_args()

    if not os.path.isdir(args.xlsx_folder):
        raise SystemExit(f"ERROR: folder not found: {args.xlsx_folder}")

    terms = read_terms(args.terms_txt)
    if not terms:
        raise SystemExit("ERROR: terms file is empty (or only comments/blank lines).")

    xlsx_paths = find_xlsx_files(args.xlsx_folder)
    if not xlsx_paths:
        raise SystemExit(f"ERROR: no .xlsx files found in folder: {args.xlsx_folder}")

    genomes: List[str] = [os.path.basename(p) for p in xlsx_paths]

    hits: Dict[str, Dict[str, List[str]]] = {g: {t: [] for t in terms} for g in genomes}
    descs: Dict[str, Dict[str, str]] = {g: {} for g in genomes}

    # Parse each genome
    for xlsx_path, genome in zip(xlsx_paths, genomes):
        term_to_gene_ids, gene_to_desc = extract_hits_from_xlsx_contains_originalDE(
            xlsx_path,
            terms,
            case_insensitive=(not args.case_sensitive),
        )

        for t in terms:
            hits[genome][t] = term_to_gene_ids.get(t, [])

        descs[genome] = gene_to_desc

        total_hits = sum(len(term_to_gene_ids.get(t, [])) for t in terms)
        print(f"[OK] {genome}: {total_hits} hits across {len(terms)} terms (searched only original_DE)")

    # Write output CSV (one file; term-by-term blocks)
    with open(args.out, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)

        for term in terms:
            # Row 1: term label
            w.writerow([term] + [""] * (2 * len(genomes) - 1))

            # Row 2: headers
            header_row: List[str] = []
            for g in genomes:
                header_row.extend([f"{g} | Gene_ID", f"{g} | Description"])
            w.writerow(header_row)

            max_len = max(len(hits[g][term]) for g in genomes)

            if max_len == 0:
                w.writerow([""] * (2 * len(genomes)))
                w.writerow([])
                continue

            for i in range(max_len):
                out_row: List[str] = []
                for g in genomes:
                    gene_list = hits[g][term]
                    if i < len(gene_list):
                        gid = gene_list[i]
                        out_row.append(gid)
                        out_row.append(descs[g].get(gid, ""))
                    else:
                        out_row.extend(["", ""])
                w.writerow(out_row)

            w.writerow([])

    print(f"\nDone. Output written to: {args.out}")


if __name__ == "__main__":
    main()
